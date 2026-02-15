import io

from openpyxl import load_workbook


def parse_xlsx(data: bytes) -> list[dict]:
    """Parse XLSX and return list of {text, metadata} per sheet."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells)
            if line.strip(" |"):
                rows.append(line)
        if rows:
            sheets.append({"text": "\n".join(rows), "metadata": {"sheet": sheet_name}})
    wb.close()
    return sheets
